import csv
import math

import openpyxl
from random import randint
from transliterate import translit
from datetime import date, datetime, timedelta
import random


class QueryGenerator:
    def __init__(self):
        self.names = []
        self.phone = []
        self.adjectives = []
        self.has_amount = {}
        self.products = []

    def read_xslx(self):
        wb = openpyxl.load_workbook("my-calend.xlsx")
        ws = wb.active
        is_first = True
        for row in ws.iter_rows(values_only=True):
            if is_first:
                is_first = False
            else:
                self.names.append(row[0])
                self.phone.append(row[1])

        return self

    def read_xslx2(self):
        wb = openpyxl.load_workbook("en2.xlsx")
        ws = wb.active

        start_date = date.today().replace(day=1, month=1).toordinal()
        end_date = date.today().replace(year=2028).toordinal()

        is_first = True
        for row in ws.iter_rows(values_only=True):
            if is_first:
                is_first = False
            elif row != "":
                self.names.append(row[0])

        return self

    def read_xslx3(self):
        wb = openpyxl.load_workbook("adjectives.xlsx")
        ws = wb.active

        start_date = date.today().replace(day=1, month=1).toordinal()
        end_date = date.today().replace(year=2028).toordinal()

        is_first = True
        for row in ws.iter_rows(values_only=True):
            if is_first:
                is_first = False
            elif row != "":
                self.adjectives.append(row[0])

        return self

    def read_xslx4(self):
        wb = openpyxl.load_workbook("bd.xlsx")
        ws = wb.active

        is_first = True
        for row in ws.iter_rows(values_only=True):
            if is_first:
                is_first = False
            elif row != "" and row[0] != None:
                variation_id, user_id = row[0][:-1].split(",")
                variation_id = int(variation_id.replace("(", ""))
                user_id = int(user_id.replace(" ", "").replace(")", ""))
                self.has_amount[user_id] = variation_id
                self.products.append(user_id)

        return self


    def generate_insert_user(self):
        sql = "insert into user"
        raw = ''
        user_id = 0
        start_date = date.today().replace(day=1, month=1).toordinal()
        end_date = date.today().replace(year=2028).toordinal()

        for _ in range(1000):
            random_day = date.fromordinal(random.randint(start_date, end_date))
            name = self.names[randint(0, len(self.names) - 1)]
            card_num = ""
            for _ in range(4):
                card_num += str(randint(1000, 9999))
            phone_num = "7"
            for _ in range(2):
                phone_num += str(randint(10000, 99999))
            translit_name = translit(name, language_code='ru', reversed=True).replace("'", "")
            raw += f"('{phone_num}', '{name}', '{card_num}', {randint(100, 999)}, '{translit_name}', '{random_day}'),\n"

            user_id += 1

        return raw

    def generate_insert_variation(self):
        raw = ''
        variation_names = ["bag", "box", "jar", "blant"]
        var2 = ["bag", "blant"]
        for id in range(1000):
            weight1 = randint(1, 100)
            if weight1 < 70:
                weight = randint(1, 8)
                name = var2[randint(0, 1)]
            elif weight1 >= 70:
                weight = randint(8, 100)
            raw += f"({weight}, {weight*randint(1, 10)}),\n"
        return raw

    def generate_insert_product(self):
        raw = ''
        variation = ["Долбит нормально", "Готов к употреблению", "Удобен крошить", "Забористый гаш",
                     "Идеален для девочек", "Уносит разом", "Для всей семьи", "Вся на Москва на этом сидит",
                     "Эта травка очень популярна  Казахстане", "Для деловых встреч", "Выбирайте эту траву, если хотите погрустить",
                     "Пацаны будут довольны", "Ни один дед инсайд не устоит перед тем, что вы можете заказать",
                     "Для дня рождения твоей бывшей", "Долгий эффект"
                     ]

        variation_names = ["пакет", "завернуты блант", "банка", "коробока"]
        for id in range(1000):
            adj = self.adjectives[randint(0, len(self.adjectives) - 1)].replace("'", "")
            names = self.names[randint(0, len(self.names) - 1)].replace("'", "")
            raw += f"('{adj} {names}', " \
                f"'{variation[randint(0, len(variation) - 1)]}. {variation[randint(0, len(variation) - 1)]}. {variation[randint(0, len(variation) - 1)]}. {variation[randint(0, len(variation) - 1)]}.', " \
                f"{randint(0, 3)}, " \
                f"{randint(-89, 89)}.{randint(100000, 999999)}, " \
                f"{randint(-179, 179)}.{randint(100000, 999999)}, " \
                f"{randint(0, 119)}, {randint(0, 2)}), \n"

        return raw

    def generate_insert_storage(self):
        raw = ''
        variation_names = ["пакет", "завернуты блант", "банка", "коробока"]
        for id, name in enumerate(variation_names):
            raw += f"({id}, '{name}'),\n"

        return raw

    def generate_insert_has_amount(self):
        raw = ''
        raw2 = ''
        variations = []
        for id in range(3500):
            cur_raw = f"{randint(0, 999), randint(0, 999)},\n"
            if cur_raw not in variations:
                variations.append(cur_raw)
                raw += cur_raw

        for _ in range(200):
            pass

        return raw

    def generate_insert_(self):
        raw = ''
        raw2 = ''
        variations = []
        for id in range(3500):
            cur_raw = f"{randint(0, 999), randint(0, 999)},\n"
            if cur_raw not in variations:
                variations.append(cur_raw)
                raw += cur_raw

        for _ in range(200):
            pass

        return raw

    def generate_insert_lives_at_and_order(self):
        raw = ''
        raw2 = ''
        lives_at = {}
        user_ids = []
        for id in range(50):
            user = randint(0, 999)
            address = randint(0, 4)
            cur_raw = f"({user}, {address}),\n"
            lives_at[user] = address
            user_ids.append(user)
            raw += cur_raw

        for id in range(100):
            is_delivered = randint(0, 1)
            status = 0
            user = user_ids[randint(0, len(user_ids) - 1)]
            start_date = date.today().replace(day=1, month=1).toordinal()
            end_date = date.today().replace(year=2022).toordinal()
            random_day = datetime.fromordinal(random.randint(start_date, end_date))
            random_day = random_day.__add__(timedelta(hours=randint(0, 23), minutes=randint(0, 59), seconds=randint(0, 59)))
            date_delivery = random_day.__add__(timedelta(days=3, hours=randint(0, 23)))
            if is_delivered:
                status = randint(0, 10)
                date_delivery = "NULL"
            else:
                status = 11
                date_delivery = f"'{date_delivery}'"
            raw2 += f"({id}, {status}, '{random_day}', {date_delivery}, {user}, {lives_at[user]}),\n"

        return raw + raw2

    def generate_insert_record(self):
        raw = ''
        variations = []
        for id in range(500):
            product_id = self.products[randint(0, len(self.products) - 1)]
            cur_raw = f"({randint(1, 4)}, {randint(0, 99)}, {product_id}, {self.has_amount[product_id]}),\n"
            raw += cur_raw

        return raw

    def generate_relationship(self):
        raw = ''
        variations = []
        for id in range(2000):
            cur_raw = f"({randint(0, 999)}, {randint(0, 9)}),\n"
            if cur_raw not in variations:
                variations.append(cur_raw)
                raw += cur_raw

        return raw

    def delete_id(self):
        wb = openpyxl.load_workbook("bd.xlsx")
        ws = wb.active
        raw = ''

        is_first = True
        for row in ws.iter_rows(values_only=True):
            if is_first:
                is_first = False
            elif row != "" and row[0] != None:
                item = row[0]
                raw += f"({item[item.find(' ') + 1:]}\n"

        return raw


if __name__ == '__main__':
    q = QueryGenerator()

    print(q.delete_id())
